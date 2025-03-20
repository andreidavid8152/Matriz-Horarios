import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta del archivo Excel
ruta_archivo = r"C:\Users\andrei.flores\Documents\Laboratorios PA\Programacion.xlsx"
df = pd.read_excel(ruta_archivo, sheet_name="Programación")

# Limpiar espacios extra al inicio y al final de los textos
df["MATERIA_TITULO_PA"] = df["MATERIA_TITULO_PA"].astype(str).str.strip()
df["TIPO_SALA_DESC"] = df["TIPO_SALA_DESC"].astype(str).str.strip()
df["SALA"] = df["SALA"].astype(str).str.strip()

# Definir las materias objetivo (las 4 que queremos evaluar)
materias_objetivo = [
    "D-ANATOM CLINIC E IMAGENOL I",
    "D-EMBRIOLOGIA",
    "D-NEUROANATOMIA",
    "D-PROPEDEUTICA CLINICA I",
]

# Lista extendida de materias (que pueden ocupar las aulas)
materias_ocupacion = [
    "D-ANATOM CLINIC E IMAGENOL I",
    "D-EMBRIOLOGIA",
    "D-NEUROANATOMIA",
    "D-PROPEDEUTICA CLINICA I",
    "D-PROCEDIMIENTOS CLINICOS I",
    "D-OFTALMOLOGIA Y ORL",
    "D-ANATOM CLINIC E IMAGENOL II",
    "D-NEUROREHABILITACION ADULTOS",
    "D-FISIO MEDIC Y LABORAT CLINIC",
    "D-PROCEDIMIENTOS BASICOS I",
    "D-FISIO MEDICA Y LABOR CLINIC",
    "D-PROCEDIMIENTOS BASICOS II",
    "D-CIRUGIA GENERAL",
    "D-ANATO Y FISIO SIST CARDIOVAS",
    "D-SEMIOLOGIA MEDICA II",
    "D-METODOS DIAGNOSTICOS",
    "D-FISIOLOGIA DEL DEPORTE",
    "D-CONSOLIDACION CIEN. BASICAS",
    "D-MORFOFUNCION II",
    "D-MORFO-FISIOLOGIA HUMANA III",
    "D-ANATOMIA Y FISIO SIST NERVIO",
    "D-ANATOY FISIO SISTEMA RESPIRA",
    "D-SEMIOLOGIA MEDICA I",
    "D-MORFOFUNCION I",
    "D-ANATOMIA DE SISTEMAS",
    "D-TERAPIA TRAUMATOLOGICA",
    "D-ANATOMIA Y FISIOLOG SIST MUS",
    "D-MEDICINA LEGAL",
    "D-TECNICAS ESPECIFICAS",
    "D-FISIOLOGIA DE SISTEMAS",
    "D-TERAPIA GERIATRICA",
    "D-NEUROREHABILITACION PEDIATR",
    "D-MORFO-FISIOLOGIA HUMANA II",
]

# Tipos de sala a filtrar
tipos_sala_filtrar = ["MORFOFUNCION O LAB. DESTREZAS", "CONSULTORIO CLINICO"]

# Filtrar para las materias objetivo (para obtener las aulas donde se imparten)
df_target = df[
    (df["MATERIA_TITULO_PA"].isin(materias_objetivo))
    & (df["TIPO_SALA_DESC"].isin(tipos_sala_filtrar))
]

# Filtrar para la ocupación (todas las materias que pueden ocupar las aulas)
df_ocupacion = df[
    (df["MATERIA_TITULO_PA"].isin(materias_ocupacion))
    & (df["TIPO_SALA_DESC"].isin(tipos_sala_filtrar))
]

# Días de la semana y mapeo para DAY_ID
dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
mapa_dias = {1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes", 6: "Sábado"}

# Generar bloques horarios desde 07:00 hasta 21:50
bloques = []
hora_inicio_col = []
hora_fin_col = []
hora_base = datetime.strptime("07:00", "%H:%M")
hora_maxima = datetime.strptime("21:50", "%H:%M")
i = 0
while True:
    # Cada bloque inicia 5 minutos más tarde que el anterior (además de 1 hora)
    hora_inicio = hora_base + timedelta(hours=i, minutes=5 * i)
    hora_fin_normal = hora_inicio + timedelta(hours=1)

    if hora_inicio >= hora_maxima:
        break

    if hora_fin_normal > hora_maxima:
        hora_fin = hora_maxima
    else:
        hora_fin = hora_fin_normal

    bloques.append(i)
    hora_inicio_col.append(hora_inicio.strftime("%H:%M"))
    hora_fin_col.append(hora_fin.strftime("%H:%M"))

    if hora_fin == hora_maxima:
        break
    i += 1

# Diccionario para guardar la matriz de cada materia objetivo
matrices_materias = {}

# Para cada materia objetivo se evaluarán las aulas donde se dicta y se calculará la ocupación
for materia in materias_objetivo:
    # Filtrar para obtener solo las sesiones de la materia objetivo (para identificar aulas)
    df_materia = df_target[df_target["MATERIA_TITULO_PA"] == materia]
    # Lista de aulas donde se imparte la materia
    aulas_objetivo = df_materia["SALA"].dropna().unique().tolist()

    # Inicializar matriz (celdas en "0") para la materia
    matriz = pd.DataFrame("0", index=bloques, columns=dias_semana)

    # Creamos un diccionario auxiliar para guardar, por bloque y día, las aulas ocupadas (para evitar contar duplicados)
    ocupacion = {(idx, dia): set() for idx in bloques for dia in dias_semana}

    # Para cada aula en las aulas donde se dicta la materia
    for aula in aulas_objetivo:
        # Filtrar las sesiones de ocupación en esa aula (usando la lista extendida)
        df_aula = df_ocupacion[df_ocupacion["SALA"] == aula]
        if df_aula.empty:
            continue

        # Procesar cada sesión en esa aula
        for _, row in df_aula.iterrows():
            # Convertir las horas a formato datetime (se asegura de tener 4 dígitos)
            try:
                inicio_str = str(row["HORA_INICIO"]).zfill(4)
                fin_str = str(row["HORA_FIN"]).zfill(4)
                hora_inicio_real = datetime.strptime(inicio_str, "%H%M")
                hora_fin_real = datetime.strptime(fin_str, "%H%M")
            except:
                continue

            # Obtener el día según DAY_ID
            dia_id = row.get("DAY_ID", None)
            if pd.isna(dia_id) or dia_id not in mapa_dias:
                continue
            dia = mapa_dias[dia_id]

            # Verificar para cada bloque si se solapa con la sesión
            for idx in range(len(hora_inicio_col)):
                inicio_bloque = datetime.strptime(hora_inicio_col[idx], "%H:%M")
                fin_bloque = datetime.strptime(hora_fin_col[idx], "%H:%M")

                # Si el bloque se solapa con la sesión, se marca el aula como ocupada
                if inicio_bloque < hora_fin_real and fin_bloque > hora_inicio_real:
                    ocupacion[(idx, dia)].add(aula)

    # Llenar la matriz con la cantidad de aulas ocupadas por bloque y día
    for idx in bloques:
        for dia in dias_semana:
            count = len(ocupacion[(idx, dia)])
            matriz.loc[idx, dia] = str(count) if count > 0 else "0"

    # Agregar columnas con la hora de inicio y fin de cada bloque
    matriz.insert(0, "HORA_INICIO", hora_inicio_col)
    matriz.insert(1, "HORA_FIN", hora_fin_col)

    matrices_materias[materia] = matriz

# Guardar cada matriz en una hoja de un archivo Excel
nombre_archivo_salida = "matrices_ocupacion_por_materia.xlsx"
with pd.ExcelWriter(nombre_archivo_salida, engine="openpyxl") as writer:
    hojas_agregadas = 0
    for materia, matriz in matrices_materias.items():
        # Limitar el nombre de la hoja a 31 caracteres y limpiar caracteres no permitidos
        hoja_nombre = (
            re.sub(r"[\\/*?:\[\]/]", "_", materia.strip())[:31] or "Materia_SinNombre"
        )
        matriz.to_excel(writer, sheet_name=hoja_nombre, index=False)
        hojas_agregadas += 1
    if hojas_agregadas == 0:
        aviso = pd.DataFrame(
            {"Mensaje": ["No se encontraron clases con las condiciones dadas."]}
        )
        aviso.to_excel(writer, sheet_name="SinDatos")

# Aplicar formato de colores con openpyxl:
# Verde para celdas con "0" (disponible) y rojo para celdas con valor mayor a 0 (ocupado)
wb = load_workbook(nombre_archivo_salida)
fill_disponible = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)  # Verde
fill_ocupado = PatternFill(
    start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
)  # Rojo

for hoja in wb.sheetnames:
    ws = wb[hoja]
    if hoja == "SinDatos":
        continue
    # Las celdas de interés empiezan en la tercera columna (después de HORA_INICIO y HORA_FIN)
    for fila in ws.iter_rows(min_row=2, min_col=3):
        for celda in fila:
            try:
                valor = int(celda.value)
            except:
                valor = 0
            if valor == 0:
                celda.fill = fill_disponible
            elif valor > 0:
                celda.fill = fill_ocupado

wb.save(nombre_archivo_salida)
print("✅ Archivo final correctamente.")
