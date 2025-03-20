import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta del archivo Excel
ruta_archivo = r"C:\Users\andrei.flores\Documents\Laboratorios PA\Programacion.xlsx"
df = pd.read_excel(ruta_archivo, sheet_name="Programación")

# Limpiar espacios extra al inicio y al final de los textos
df["MATERIA_TITULO_PA"] = df["MATERIA_TITULO_PA"].astype(str).str.strip()
df["TIPO_SALA_DESC"] = df["TIPO_SALA_DESC"].astype(str).str.strip()
df["SALA"] = df["SALA"].astype(str).str.strip()

# Materias objetivo
materias_objetivo = [
    "D-ANATOM CLINIC E IMAGENOL I",
    "D-EMBRIOLOGIA",
    "D-NEUROANATOMIA",
    "D-PROPEDEUTICA CLINICA I",
]

# Lista extendida de materias
materias_ocupacion = [
    "D-ANATOM CLINIC E IMAGENOL I",
    "D-EMBRIOLOGIA",
    "D-NEUROANATOMIA",
    "D-PROPEDEUTICA CLINICA I",
    "D-PROCEDIMIENTOS CLINICOS I",
    "D-OFTALMOLOGIA Y ORL",
    "D-ANATOM CLINIC E IMAGENOL II",
    "D-NEUROREHABILITACION ADULTOS",
    "D-FISIO MEDIC Y LABORAT CLINIC",
    "D-PROCEDIMIENTOS BASICOS I",
    "D-FISIO MEDICA Y LABOR CLINIC",
    "D-PROCEDIMIENTOS BASICOS II",
    "D-CIRUGIA GENERAL",
    "D-ANATO Y FISIO SIST CARDIOVAS",
    "D-SEMIOLOGIA MEDICA II",
    "D-METODOS DIAGNOSTICOS",
    "D-FISIOLOGIA DEL DEPORTE",
    "D-CONSOLIDACION CIEN. BASICAS",
    "D-MORFOFUNCION II",
    "D-MORFO-FISIOLOGIA HUMANA III",
    "D-ANATOMIA Y FISIO SIST NERVIO",
    "D-ANATOY FISIO SISTEMA RESPIRA",
    "D-SEMIOLOGIA MEDICA I",
    "D-MORFOFUNCION I",
    "D-ANATOMIA DE SISTEMAS",
    "D-TERAPIA TRAUMATOLOGICA",
    "D-ANATOMIA Y FISIOLOG SIST MUS",
    "D-MEDICINA LEGAL",
    "D-TECNICAS ESPECIFICAS",
    "D-FISIOLOGIA DE SISTEMAS",
    "D-TERAPIA GERIATRICA",
    "D-NEUROREHABILITACION PEDIATR",
    "D-MORFO-FISIOLOGIA HUMANA II",
]

# Tipos de sala a filtrar
tipos_sala_filtrar = ["MORFOFUNCION O LAB. DESTREZAS", "CONSULTORIO CLINICO"]

# Primer filtro: materias objetivo
df_target = df[
    (df["MATERIA_TITULO_PA"].isin(materias_objetivo))
    & (df["TIPO_SALA_DESC"].isin(tipos_sala_filtrar))
]

# Segundo filtro: materias de ocupación
df_ocupacion = df[
    (df["MATERIA_TITULO_PA"].isin(materias_ocupacion))
    & (df["TIPO_SALA_DESC"].isin(tipos_sala_filtrar))
]

# Días y mapeo
dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
mapa_dias = {1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes", 6: "Sábado"}

# Crear bloques horarios
bloques = []
hora_inicio_col = []
hora_fin_col = []
hora_base = datetime.strptime("07:00", "%H:%M")
hora_maxima = datetime.strptime("21:50", "%H:%M")
i = 0
while True:
    hora_inicio = hora_base + timedelta(hours=i, minutes=5 * i)
    hora_fin_normal = hora_inicio + timedelta(hours=1)
    if hora_inicio >= hora_maxima:
        break
    hora_fin = hora_maxima if hora_fin_normal > hora_maxima else hora_fin_normal
    bloques.append(i)
    hora_inicio_col.append(hora_inicio.strftime("%H:%M"))
    hora_fin_col.append(hora_fin.strftime("%H:%M"))
    if hora_fin == hora_maxima:
        break
    i += 1

# Diccionario para matrices
matrices_materias = {}

for materia in materias_objetivo:
    df_materia = df_target[df_target["MATERIA_TITULO_PA"] == materia]
    aulas_objetivo = df_materia["SALA"].dropna().unique().tolist()

    matriz = pd.DataFrame("0", index=bloques, columns=dias_semana)

    ocupacion = {(idx, dia): set() for idx in bloques for dia in dias_semana}
    aulas_totales = set()

    for aula in aulas_objetivo:
        df_aula = df_ocupacion[df_ocupacion["SALA"] == aula]
        if df_aula.empty:
            continue
        for _, row in df_aula.iterrows():
            try:
                inicio_str = str(row["HORA_INICIO"]).zfill(4)
                fin_str = str(row["HORA_FIN"]).zfill(4)
                hora_inicio_real = datetime.strptime(inicio_str, "%H%M")
                hora_fin_real = datetime.strptime(fin_str, "%H%M")
            except:
                continue
            dia_id = row.get("DAY_ID", None)
            if pd.isna(dia_id) or dia_id not in mapa_dias:
                continue
            dia = mapa_dias[dia_id]
            for idx in range(len(hora_inicio_col)):
                inicio_bloque = datetime.strptime(hora_inicio_col[idx], "%H:%M")
                fin_bloque = datetime.strptime(hora_fin_col[idx], "%H:%M")
                if inicio_bloque < hora_fin_real and fin_bloque > hora_inicio_real:
                    ocupacion[(idx, dia)].add(aula)
                    aulas_totales.add(aula)

    for idx in bloques:
        for dia in dias_semana:
            aulas_ocup = sorted(ocupacion[(idx, dia)])
            count = len(aulas_ocup)
            if count > 0:
                matriz.loc[idx, dia] = f"{count}: {', '.join(aulas_ocup)}"
            else:
                matriz.loc[idx, dia] = "0"

    matriz.insert(0, "HORA_INICIO", hora_inicio_col)
    matriz.insert(1, "HORA_FIN", hora_fin_col)
    matriz[""] = ""  # columna vacía 1
    matriz["  "] = ""  # columna vacía 2

    # Añadir columna con aulas ocupadas (solo en la primera fila)
    columna_aulas = [""] * len(matriz)
    columna_aulas[0] = "Aulas ocupadas: " + (
        ", ".join(sorted(aulas_totales)) if aulas_totales else "Ninguna"
    )
    matriz["AULAS_MATERIA"] = columna_aulas

    matrices_materias[materia] = matriz

# Guardar archivo Excel
nombre_archivo_salida = "matrices_ocupacion_por_materia_detalle.xlsx"
with pd.ExcelWriter(nombre_archivo_salida, engine="openpyxl") as writer:
    hojas_agregadas = 0
    for materia, matriz in matrices_materias.items():
        hoja_nombre = (
            re.sub(r"[\\/*?:\[\]/]", "_", materia.strip())[:31] or "Materia_SinNombre"
        )
        matriz.to_excel(writer, sheet_name=hoja_nombre, index=False)
        hojas_agregadas += 1
    if hojas_agregadas == 0:
        aviso = pd.DataFrame(
            {"Mensaje": ["No se encontraron clases con las condiciones dadas."]}
        )
        aviso.to_excel(writer, sheet_name="SinDatos")

# Aplicar formato de colores
wb = load_workbook(nombre_archivo_salida)
fill_disponible = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
fill_ocupado = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

for hoja in wb.sheetnames:
    if hoja == "SinDatos":
        continue
    ws = wb[hoja]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=7):
        for celda in row:
            try:
                valor = celda.value
                if isinstance(valor, str) and ":" in valor:
                    num = int(valor.split(":")[0].strip())
                else:
                    num = int(valor)
            except:
                num = 0
            if num == 0:
                celda.fill = fill_disponible
            elif num > 0:
                celda.fill = fill_ocupado

wb.save(nombre_archivo_salida)
print("✅ Archivo final creado correctamente.")
