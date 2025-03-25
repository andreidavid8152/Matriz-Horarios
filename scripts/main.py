import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Cargar archivo Excel
ruta_archivo = r"C:\Users\andrei.flores\Documents\Laboratorios PA\Programacion.xlsx"
df = pd.read_excel(ruta_archivo, sheet_name="Programación")

# Materias y tipos de sala
materias_filtrar = [
    "D-ANATOM CLINIC E IMAGENOL I",
    "D-EMBRIOLOGIA",
    "D-NEUROANATOMIA",
    "D-PROPEDEUTICA CLINICA I",
    "D-PROCEDIMIENTOS CLINICOS I",
    "D-OFTALMOLOGIA Y ORL",
    "D-ANATOM CLINIC E IMAGENOL II",
    "D-NEUROREHABILITACION ADULTOS",
    "D-FISIO MEDIC Y LABORAT CLINIC",
    "D-PROCEDIMIENTOS BASICOS I",
    "D-FISIO MEDICA Y LABOR CLINIC",
    "D-PROCEDIMIENTOS BASICOS II",
    "D-CIRUGIA GENERAL",
    "D-ANATO Y FISIO SIST CARDIOVAS",
    "D-SEMIOLOGIA MEDICA II",
    "D-METODOS DIAGNOSTICOS",
    "D-FISIOLOGIA DEL DEPORTE",
    "D-CONSOLIDACION CIEN. BASICAS",
    "D-MORFOFUNCION II",
    "D-MORFO-FISIOLOGIA HUMANA III",
    "D-ANATOMIA Y FISIO SIST NERVIO",
    "D-ANATOY FISIO SISTEMA RESPIRA",
    "D-SEMIOLOGIA MEDICA I",
    "D-MORFOFUNCION I",
    "D-ANATOMIA DE SISTEMAS",
    "D-TERAPIA TRAUMATOLOGICA",
    "D-ANATOMIA Y FISIOLOG SIST MUS",
    "D-MEDICINA LEGAL",
    "D-TECNICAS ESPECIFICAS",
    "D-FISIOLOGIA DE SISTEMAS",
    "D-TERAPIA GERIATRICA",
    "D-NEUROREHABILITACION PEDIATR",
    "D-MORFO-FISIOLOGIA HUMANA II",
]


tipos_sala_filtrar = ["MORFOFUNCION O LAB. DESTREZAS", "CONSULTORIO CLINICO"]

# Filtrar DataFrame
df_filtrado = df[
    (df["MATERIA_TITULO_PA"].isin(materias_filtrar))
    & (df["TIPO_SALA_DESC"].isin(tipos_sala_filtrar))
]

# Aulas únicas
aulas_unicas = df_filtrado["SALA"].dropna().unique().tolist()

# Días de la semana y mapeo
dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
mapa_dias = {1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes"}

# Generar bloques dinámicamente hasta que la hora_fin sea 21:50
bloques = []
hora_base = datetime.strptime("07:00", "%H:%M")
hora_maxima = datetime.strptime("21:50", "%H:%M")
hora_inicio_col = []
hora_fin_col = []

i = 0
while True:
    # Hora de inicio de cada bloque (1 hora + 5 min más tarde que el anterior)
    hora_inicio = hora_base + timedelta(hours=i, minutes=5 * i)
    # Fin normal del bloque
    hora_fin_normal = hora_inicio + timedelta(hours=1)

    # Si la hora de inicio ya es mayor o igual a 21:50, no creamos más bloques
    if hora_inicio >= hora_maxima:
        break

    # Si la hora de fin normal sobrepasa 21:50, creamos un bloque parcial que termina a 21:50
    if hora_fin_normal > hora_maxima:
        hora_fin = hora_maxima
    else:
        hora_fin = hora_fin_normal

    # Agregamos este bloque
    bloques.append(i)
    hora_inicio_col.append(hora_inicio.strftime("%H:%M"))
    hora_fin_col.append(hora_fin.strftime("%H:%M"))

    # Si se ajustó a 21:50, salimos porque no hay más tiempo
    if hora_fin == hora_maxima:
        break

    i += 1

# Crear matrices
matrices_ocupacion = {}

for aula in aulas_unicas:
    matriz = pd.DataFrame("0", index=bloques, columns=dias_semana)
    df_aula = df_filtrado[df_filtrado["SALA"] == aula]

    if df_aula.empty:
        continue

    for dia in dias_semana:
        # Tomar el día ID (1=Lunes, 2=Martes, etc.)
        dia_id = list(mapa_dias.keys())[list(mapa_dias.values()).index(dia)]
        df_dia = df_aula[df_aula["DAY_ID"] == dia_id]

        if df_dia.empty:
            continue

        # Agrupar por NRC y MATERIA
        grupos = df_dia.groupby(["NRC", "MATERIA_TITULO_PA"])

        for (nrc, materia), grupo in grupos:
            try:
                # Tomamos la hora más temprana y la más tardía para esa combinación NRC+Materia
                min_inicio = grupo["HORA_INICIO"].astype(str).str.zfill(4).min()
                max_fin = grupo["HORA_FIN"].astype(str).str.zfill(4).max()

                hora_inicio_real = datetime.strptime(min_inicio, "%H%M")
                hora_fin_real = datetime.strptime(max_fin, "%H%M")
            except:
                continue

            # Marcar todos los bloques que se solapan con el rango real
            for idx, _ in enumerate(hora_inicio_col):
                inicio_bloque = datetime.strptime(hora_inicio_col[idx], "%H:%M")
                fin_bloque = datetime.strptime(hora_fin_col[idx], "%H:%M")

                # Si este bloque se solapa con la clase real, lo marcamos como ocupado
                if inicio_bloque < hora_fin_real and fin_bloque > hora_inicio_real:
                    matriz.loc[idx, dia] = f"1 - {materia}"

    # Agregar columnas de hora (Inicio/Fin)
    matriz.insert(0, "HORA_INICIO", hora_inicio_col)
    matriz.insert(1, "HORA_FIN", hora_fin_col)

    matrices_ocupacion[aula] = matriz

# Guardar archivo Excel
nombre_archivo_salida = "matrices_ocupacion_por_aula.xlsx"
with pd.ExcelWriter(nombre_archivo_salida, engine="openpyxl") as writer:
    hojas_agregadas = 0

    for aula, matriz in matrices_ocupacion.items():
        hoja_nombre = (
            re.sub(r"[\\/*?:\[\]/]", "_", str(aula).strip())[:31] or "Aula_SinNombre"
        )
        matriz.to_excel(writer, sheet_name=hoja_nombre, index=False)
        hojas_agregadas += 1

    if hojas_agregadas == 0:
        aviso = pd.DataFrame(
            {"Mensaje": ["No se encontraron clases con las condiciones dadas."]}
        )
        aviso.to_excel(writer, sheet_name="SinDatos")

# Aplicar formato de colores
wb = load_workbook(nombre_archivo_salida)
fill_disponible = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)  # Verde
fill_ocupado = PatternFill(
    start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
)  # Rojo

for hoja in wb.sheetnames:
    ws = wb[hoja]
    if hoja == "SinDatos":
        continue

    for fila in ws.iter_rows(min_row=2, min_col=3):
        for celda in fila:
            if str(celda.value).strip() == "0":
                celda.fill = fill_disponible
            elif isinstance(celda.value, str) and celda.value.startswith("1 -"):
                celda.fill = fill_ocupado

wb.save(nombre_archivo_salida)

print(
    "✅ Archivo final correctamente."
)
